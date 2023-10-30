from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from subprocess import CREATE_NO_WINDOW
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
global driver
import time

import sys
print("Python Version: " + str(sys.version_info))
service = Service()
service.creationflags = CREATE_NO_WINDOW
opts = Options()
opts.binary_location = 'C:\\Program Files\\Mozilla Firefox\\firefox.exe'
opts.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv")

def scroll():
    for i in range(100):
        script = f"window.scrollBy(0, {500});"
        driver.execute_script(script)
        time.sleep(0.1)

def links_ciudad():
    driver.get(
        "https://www.linkedin.com/jobs/search?keywords=ingeniero%20de%20sistemas&location=Bogota%2C%20D.C.&geoId=&trk=people-guest_jobs-search-bar_search-submit&position=1&pageNum=0")
    scroll()
    elements = driver.find_elements("xpath", '//*[@id="main-content"]/section[2]/ul/li[*]/div/a')
    for element in elements:
        href = element.get_attribute("href")
        with open("Links/links_ciudad.txt", "a") as f:
            f.write(href + '\n')

def ciudad():
    archivo = open('Links/links_ciudad.txt', 'r')
    wb = Workbook()
    hoja_activa = wb.active
    fila = 1
    driver.get("https://www.linkedin.com/login")
    time.sleep(10)
    for link in archivo:
        try:
            driver.get(link)
            time.sleep(0.5)

            celda_titulo = 'A' + str(fila)
            titulo_elemento =  WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[1]/div/div/div[1]/div[1]/h1'))
            )
            titulo = titulo_elemento.text
            hoja_activa[celda_titulo] = titulo

            celda_empresa = 'B'+str(fila)
            empresa_elemento = driver.find_element("xpath", '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[1]/div/div/div[1]/div[2]/div')
            empresa = empresa_elemento.text
            hoja_activa[celda_empresa] = empresa

            celda_ubicacion = 'C' + str(fila)
            hoja_activa[celda_ubicacion] = link

            boton = driver.find_element("xpath", '//*[@id="ember37"]')
            boton.click()

            celda_contenido = 'D' + str(fila)
            contenido_elemento = driver.find_element("xpath", '//*[@id="job-details"]')
            contenido = contenido_elemento.text
            hoja_activa[celda_contenido] = contenido

            aptitudes_link = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="how-you-match-card-container"]/section[2]/div/div/a'))
            )
            aptitudes_link.click()

            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[3]/div/div/div[2]/div/div[1]/ul/li[1]/div/div[2]'))
            )

            celda_aptitudes = 'E' + str(fila)
            aptitudes = driver.find_elements("xpath", '/html/body/div[3]/div/div/div[2]/div/div[1]/ul/li[*]/div/div[2]')
            list_aptitudes = ''
            for aptitud in aptitudes:
                    list_aptitudes = list_aptitudes + str(aptitud.text) + '.'
            print('aptitudes: ' + str(list_aptitudes))
            hoja_activa[celda_aptitudes] = list_aptitudes

            """""
            celda_criterios = 'E' + str(fila)
            criterios_elementos_t = driver.find_elements("xpath", '/html/body/div[1]/div/section/div[2]/div/section[1]/div/ul/li')
            criterios_elementos_c = driver.find_elements("xpath", '/html/body/div[1]/div/section/div[2]/div/section[1]/div/ul/li')
            criterios = ''
            num = 1
            for i in criterios_elementos_t:
                criterios = criterios + str(criterios_elementos_t[num]) + " : " + str(criterios_elementos_c[num])
                num = num + 1
            hoja_activa[celda_criterios] = criterios
            """""
            fila = fila + 1
            time.sleep(1)
        except Exception as e:
            print("ERROR: " + str(e))
    wb.save('ciudad.xlsx')


def links_pais():
    driver.get(
        "https://www.linkedin.com/jobs/search?keywords=Ingeniero%20De%20Sistemas&location=Colombia&geoId=&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0")
    scroll()
    elements = driver.find_elements("xpath", '//*[@id="main-content"]/section[2]/ul/li[*]/div/a')
    for element in elements:
        href = element.get_attribute("href")
        with open("Links/links_pais.txt", "a") as f:
            f.write(href + '\n')

def pais():
    archivo = open('Links/links_pais.txt', 'r')
    wb = Workbook()
    hoja_activa = wb.active
    fila = 1
    driver.get("https://www.linkedin.com/login")
    time.sleep(10)
    for link in archivo:
        try:
            driver.get(link)
            time.sleep(0.5)

            celda_titulo = 'A' + str(fila)
            titulo_elemento = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH,
                                                '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[1]/div/div/div[1]/div[1]/h1'))
            )
            titulo = titulo_elemento.text
            hoja_activa[celda_titulo] = titulo

            celda_empresa = 'B' + str(fila)
            empresa_elemento = driver.find_element("xpath",
                                                   '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[1]/div/div/div[1]/div[2]/div')
            empresa = empresa_elemento.text
            hoja_activa[celda_empresa] = empresa

            celda_ubicacion = 'C' + str(fila)
            hoja_activa[celda_ubicacion] = link

            boton = driver.find_element("xpath", '//*[@id="ember37"]')
            boton.click()

            celda_contenido = 'D' + str(fila)
            contenido_elemento = driver.find_element("xpath", '//*[@id="job-details"]')
            contenido = contenido_elemento.text
            hoja_activa[celda_contenido] = contenido

            aptitudes_link = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="how-you-match-card-container"]/section[2]/div/div/a'))
            )
            aptitudes_link.click()

            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[3]/div/div/div[2]/div/div[1]/ul/li[1]/div/div[2]'))
            )

            celda_aptitudes = 'E' + str(fila)
            aptitudes = driver.find_elements("xpath", '/html/body/div[3]/div/div/div[2]/div/div[1]/ul/li[*]/div/div[2]')
            list_aptitudes = ''
            for aptitud in aptitudes:
                list_aptitudes = list_aptitudes + str(aptitud.text) + '.'
            print('aptitudes: ' + str(list_aptitudes))
            hoja_activa[celda_aptitudes] = list_aptitudes

            """""
            celda_criterios = 'E' + str(fila)
            criterios_elementos_t = driver.find_elements("xpath", '/html/body/div[1]/div/section/div[2]/div/section[1]/div/ul/li')
            criterios_elementos_c = driver.find_elements("xpath", '/html/body/div[1]/div/section/div[2]/div/section[1]/div/ul/li')
            criterios = ''
            num = 1
            for i in criterios_elementos_t:
                criterios = criterios + str(criterios_elementos_t[num]) + " : " + str(criterios_elementos_c[num])
                num = num + 1
            hoja_activa[celda_criterios] = criterios
            """""
            fila = fila + 1
            time.sleep(1)
        except Exception as e:
            print("ERROR: " + str(e))
    wb.save('pais.xlsx')

def links_mundo():
    driver.get(
        "https://www.linkedin.com/jobs/search?keywords=System%20Engineer&location=United%20States&geoId=103644278&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0")
    scroll()
    elements_eu = driver.find_elements("xpath", '//*[@id="main-content"]/section[2]/ul/li[*]/div/a')
    for index, element in enumerate(elements_eu):
        if index >= 50:
            break
        href = element.get_attribute("href")
        with open("Links/links_mundo.txt", "a") as f:
            f.write(href + '\n')

    driver.get(
        "https://www.linkedin.com/jobs/search?keywords=Ingeniero%20De%20Sistemas&location=Spain&geoId=105646813&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0")
    scroll()
    elements_sp = driver.find_elements("xpath", '//*[@id="main-content"]/section[2]/ul/li[*]/div/a')
    for index, element in enumerate(elements_sp):
        if index >= 50:
            break
        href = element.get_attribute("href")
        with open("Links/links_mundo.txt", "a") as f:
            f.write(href + '\n')

    driver.get(
        "https://www.linkedin.com/jobs/search?keywords=System%20Engineer&location=United%20Kingdom&geoId=101165590&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0")
    scroll()
    elements_uk = driver.find_elements("xpath", '//*[@id="main-content"]/section[2]/ul/li[*]/div/a')
    for index, element in enumerate(elements_uk):
        if index >= 50:
            break
        href = element.get_attribute("href")
        with open("Links/links_mundo.txt", "a") as f:
            f.write(href + '\n')

def mundo():
    archivo = open('Links/links_mundo.txt', 'r')
    wb = Workbook()
    hoja_activa = wb.active
    fila = 1
    driver.get("https://www.linkedin.com/login")
    time.sleep(10)
    for link in archivo:
        try:
            driver.get(link)
            time.sleep(0.5)

            celda_titulo = 'A' + str(fila)
            titulo_elemento = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH,
                                                '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[1]/div/div/div[1]/div[1]/h1'))
            )
            titulo = titulo_elemento.text
            hoja_activa[celda_titulo] = titulo

            celda_empresa = 'B' + str(fila)
            empresa_elemento = driver.find_element("xpath",
                                                   '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[1]/div/div/div[1]/div[2]/div')
            empresa = empresa_elemento.text
            hoja_activa[celda_empresa] = empresa

            celda_ubicacion = 'C' + str(fila)
            hoja_activa[celda_ubicacion] = link

            boton = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[3]/div/div[2]/div/div/main/div/div[1]/div/div[2]/footer/button')))
            boton.click()

            celda_contenido = 'D' + str(fila)
            contenido_elemento = driver.find_element("xpath", '//*[@id="job-details"]')
            contenido = contenido_elemento.text
            hoja_activa[celda_contenido] = contenido

            aptitudes_link = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="how-you-match-card-container"]/section[2]/div/div/a'))
            )
            aptitudes_link.click()

            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[3]/div/div/div[2]/div/div[1]/ul/li[1]/div/div[2]'))
            )

            celda_aptitudes = 'E' + str(fila)
            aptitudes = driver.find_elements("xpath", '/html/body/div[3]/div/div/div[2]/div/div[1]/ul/li[*]/div/div[2]')
            list_aptitudes = ''
            for aptitud in aptitudes:
                list_aptitudes = list_aptitudes + str(aptitud.text) + '.'
            print('aptitudes: ' + str(list_aptitudes))
            hoja_activa[celda_aptitudes] = list_aptitudes

            """""
            celda_criterios = 'E' + str(fila)
            criterios_elementos_t = driver.find_elements("xpath", '/html/body/div[1]/div/section/div[2]/div/section[1]/div/ul/li')
            criterios_elementos_c = driver.find_elements("xpath", '/html/body/div[1]/div/section/div[2]/div/section[1]/div/ul/li')
            criterios = ''
            num = 1
            for i in criterios_elementos_t:
                criterios = criterios + str(criterios_elementos_t[num]) + " : " + str(criterios_elementos_c[num])
                num = num + 1
            hoja_activa[celda_criterios] = criterios
            """""
            fila = fila + 1
            time.sleep(1)
        except Exception as e:
            print("ERROR: " + str(e))
    wb.save('mundo.xlsx')

try:
    driver = webdriver.Firefox(service=service, options=opts)
    """""
    extension_path = os.path.abspath("extension/adblock.xpi")
    driver.install_addon(extension_path, temporary=True)

    time.sleep(2)

    driver.switch_to.window(driver.window_handles[0])
    """""

    #links_ciudad()
    #ciudad()
    #links_pais()
    #pais()
    #links_mundo()
    mundo()
    driver.quit()
except Exception:
    e = sys.exc_info()[1]
    print(e)

